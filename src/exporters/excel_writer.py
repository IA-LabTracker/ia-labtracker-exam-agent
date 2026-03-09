from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.aggregator.models import INSTITUTION_ALIASES, INSTITUTIONS, ReconciledRow, ReverseRow
from src.normalize.normalizer import classify_color
from src.utils.logging import logger
from src.utils.manchester_order import lookup_semana, format_semana, SEMANA_UNKNOWN, POS_UNKNOWN

# Column key for each institution: "q_FAMERP", "q_USP", etc.
_INST_COLS = [f"q_{inst}" for inst in INSTITUTIONS]

COLUMNS = [
    "semana",
    "match_label",
    "match_score",
    "input_tema",
    "input_equivalencia",
    "normalized_tema",
    "normalized_subtema",
    *_INST_COLS,
    "notes",
]

COLUMN_HEADERS = {
    "semana": "Semana",
    "input_tema": "Tema (entrada)",
    "input_equivalencia": "Equivalência",
    "normalized_tema": "Tema",
    "normalized_subtema": "Subtema",
    **{f"q_{inst}": inst for inst in INSTITUTIONS},
    "match_label": "Confiança",
    "match_score": "Score",
    "notes": "Observações",
}

REVERSE_COLUMNS = [
    "db_tema",
    "db_subtema",
    "db_num_questions",
    "matched_input",
    "similarity_score",
    "coverage_status",
    "notes",
]

REVERSE_HEADERS = {
    "db_tema": "Tema (banco)",
    "db_subtema": "Subtema (banco)",
    "db_num_questions": "Qtd. Questões",
    "matched_input": "Entrada Correspondente",
    "similarity_score": "Similaridade",
    "coverage_status": "Status Cobertura",
    "notes": "Observações",
}

ROW_COLORS = {
    "#EF4444": "00FECACA",  # vermelho → fundo rosa claro
    "#F97316": "00FED7AA",  # laranja → fundo pêssego
    "#EAB308": "00FEF9C3",  # amarelo → fundo amarelo claro
    "#22C55E": "00DCFCE7",  # verde → fundo verde claro
    "#3B82F6": "00DBEAFE",  # azul → fundo azul claro
}

BADGE_COLORS = {
    "#EF4444": ("00EF4444", "00FFFFFF"),  # vermelho
    "#F97316": ("00F97316", "00FFFFFF"),  # laranja
    "#EAB308": ("00EAB308", "00000000"),  # amarelo
    "#22C55E": ("0022C55E", "00000000"),  # verde
    "#3B82F6": ("003B82F6", "00FFFFFF"),  # azul
}

COVERAGE_COLORS = {
    "coberto": "00DCFCE7",  # green
    "parcial": "00FEF9C3",  # yellow
    "não coberto": "00FECACA",  # red
}

CONFIDENCE_FILLS = {
    "Quente": "00DCFCE7",  # green
    "Morno": "00FEF9C3",  # yellow
    "Frio": "00DBEAFE",  # light blue
    "Sem match": "00FECACA",  # red
}

# Pre-built immutable style objects (avoid copy() per cell)
_BORDER = Border(
    left=Side(style="thin", color="00000000"),
    right=Side(style="thin", color="00000000"),
    top=Side(style="thin", color="00000000"),
    bottom=Side(style="thin", color="00000000"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_HEADER_FILL = PatternFill(
    start_color="001F1F1F", end_color="001F1F1F", fill_type="solid"
)
_HEADER_FONT = Font(color="00FFFFFF", bold=True, size=11)
_DATA_FONT = Font(size=11, color="00000000")
_BOLD_FONT = Font(size=11, bold=True, color="00000000")
_ROW_FILL = PatternFill(start_color="00EFF6FF", end_color="00EFF6FF", fill_type="solid")

# Pre-build fill objects for each color (avoid creating per cell)
_FILL_CACHE: dict[str, PatternFill] = {}

# Pre-build badge fonts (avoid creating per row)
_BADGE_FONT_CACHE: dict[str, Font] = {
    hex_val: Font(color=fg, bold=True, size=11)
    for hex_val, (bg, fg) in BADGE_COLORS.items()
}


def _get_fill(argb: str) -> PatternFill:
    """Get or create a cached PatternFill for the given ARGB color."""
    if argb not in _FILL_CACHE:
        _FILL_CACHE[argb] = PatternFill(
            start_color=argb, end_color=argb, fill_type="solid"
        )
    return _FILL_CACHE[argb]


def write_excel(
    rows: list[ReconciledRow],
    output_path: str | Path,
    also_csv: bool = False,
    reverse_rows: list[ReverseRow] | None = None,
) -> Path:
    output_path = Path(output_path)

    # Build (record, cor_hex, qbi, sort_key) tuples so we can sort by Manchester order
    raw_entries: list[tuple[dict, str, dict, tuple[int, int]]] = []
    # Pre-build case-insensitive lookup: lower(inst) → canonical inst name.
    # Includes INSTITUTION_ALIASES so short DB names (e.g. "AMP") resolve to
    # the canonical INSTITUTIONS entry (e.g. "AMP-PR").
    _inst_lower: dict[str, str] = {inst.lower(): inst for inst in INSTITUTIONS}
    for alias, canonical in INSTITUTION_ALIASES.items():
        _inst_lower[alias.lower()] = canonical

    for r in rows:
        d = asdict(r)
        cor_hex = d.pop("cor_hex", "#22C55E")
        d.pop("match_method", None)
        qbi_raw: dict[str, int] = d.pop("questions_by_institution", {})

        # Normalize institution names from DB to canonical INSTITUTIONS list
        # (handles case differences and abbreviations via INSTITUTION_ALIASES)
        qbi: dict[str, int] = {}
        for db_name, count in qbi_raw.items():
            canonical = _inst_lower.get(db_name.lower(), db_name)
            qbi[canonical] = qbi.get(canonical, 0) + count

        for inst in INSTITUTIONS:
            d[f"q_{inst}"] = f"{qbi.get(inst, 0)} questões"
        d["match_score"] = f"{d.get('match_score', 0):.0%}"

        # Derive input tema/subtema for Manchester lookup
        input_parts = (d.get("input_tema") or "").split(" | ", 1)
        input_tema_raw = input_parts[0].strip()
        input_subtema_raw = input_parts[1].strip() if len(input_parts) > 1 else None

        sort_key = lookup_semana(
            d.get("normalized_tema"),
            d.get("normalized_subtema"),
            input_tema_raw,
            input_subtema_raw,
        )
        semana_int = sort_key[0]
        d["semana"] = format_semana(semana_int)

        raw_entries.append((d, cor_hex, qbi, sort_key))

    # Sort by (semana, row_position) — rows without a Manchester match go last
    raw_entries.sort(key=lambda e: e[3])

    records = [e[0] for e in raw_entries]
    cor_hex_list = [e[1] for e in raw_entries]
    inst_counts_list = [e[2] for e in raw_entries]

    df = pd.DataFrame(records)
    df = df[[c for c in COLUMNS if c in df.columns]]

    # Write data AND style in a single pass (no load_workbook round-trip)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ranking")

        if reverse_rows:
            rev_records = []
            for rr in reverse_rows:
                d = asdict(rr)
                d.pop("db_cor_hex", None)
                d["db_num_questions"] = f"{d['db_num_questions']} questões"
                d["similarity_score"] = f"{d['similarity_score']:.2%}"
                rev_records.append(d)
            df_rev = pd.DataFrame(rev_records)
            df_rev = df_rev[[c for c in REVERSE_COLUMNS if c in df_rev.columns]]
            df_rev.to_excel(writer, index=False, sheet_name="Cobertura Reversa")

        # Style directly on the writer's workbook (avoids save → load → save)
        wb = writer.book
        _style_ranking_sheet(wb["Ranking"], df, cor_hex_list, inst_counts_list)
        if reverse_rows and "Cobertura Reversa" in wb.sheetnames:
            _style_reverse_sheet(wb["Cobertura Reversa"], reverse_rows)

    logger.info("Wrote %d rows to %s", len(df), output_path)

    if also_csv:
        csv_path = output_path.with_suffix(".csv")
        df.to_csv(csv_path, index=False)
        logger.info("Also wrote CSV: %s", csv_path)

    return output_path


def _style_ranking_sheet(
    ws,
    df: pd.DataFrame,
    cor_hex_list: list[str],
    inst_counts_list: list[dict[str, int]],
) -> None:
    num_cols = ws.max_column
    num_rows = ws.max_row

    # Track column widths during styling (eliminates second pass)
    col_max_len = [0] * (num_cols + 1)

    col_names = list(df.columns)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        col_name = col_names[col_idx - 1] if col_idx <= len(col_names) else ""
        cell.value = COLUMN_HEADERS.get(col_name, cell.value)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        if cell.value:
            col_max_len[col_idx] = len(str(cell.value))

    # Map institution column key → col_index, and reverse for O(1) per-cell lookup
    inst_col_indices: dict[str, int] = {}
    for inst in INSTITUTIONS:
        col_key = f"q_{inst}"
        if col_key in col_names:
            inst_col_indices[inst] = col_names.index(col_key) + 1
    # Reversed: col_idx → inst — avoids O(27) linear scan inside the cell loop
    col_idx_to_inst: dict[int, str] = {cidx: inst for inst, cidx in inst_col_indices.items()}

    confidence_col = (
        (col_names.index("match_label") + 1) if "match_label" in col_names else None
    )
    score_col = (
        (col_names.index("match_score") + 1) if "match_score" in col_names else None
    )

    for row_idx in range(2, num_rows + 1):
        data_idx = row_idx - 2
        qbi = inst_counts_list[data_idx] if data_idx < len(inst_counts_list) else {}

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.font = _DATA_FONT

            # Track column width
            if cell.value:
                col_max_len[col_idx] = max(col_max_len[col_idx], len(str(cell.value)))

            # Institution question cells keep their colored badges
            inst_for_col = col_idx_to_inst.get(col_idx)
            if inst_for_col is not None:
                count = qbi.get(inst_for_col, 0)
                _, inst_hex = classify_color(count)
                inst_badge = BADGE_COLORS.get(inst_hex)
                if inst_badge:
                    cell.fill = _get_fill(inst_badge[0])
                    cell.font = _BADGE_FONT_CACHE.get(inst_hex, _BOLD_FONT)
            else:
                # All other cells get a uniform light blue for readability
                cell.fill = _ROW_FILL
                if col_idx == confidence_col or col_idx == score_col:
                    cell.font = _BOLD_FONT

    ws.row_dimensions[1].height = 30
    for row_idx in range(2, num_rows + 1):
        ws.row_dimensions[row_idx].height = 25

    # Apply column widths (already computed during loop)
    for col_idx in range(1, num_cols + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(
            max(col_max_len[col_idx] + 4, 15), 55
        )


def _style_reverse_sheet(ws, reverse_rows: list[ReverseRow]) -> None:
    num_cols = ws.max_column
    num_rows = ws.max_row

    # Track column widths during styling
    col_max_len = [0] * (num_cols + 1)

    rev_col_names = list(REVERSE_COLUMNS)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        col_name = rev_col_names[col_idx - 1] if col_idx <= len(rev_col_names) else ""
        cell.value = REVERSE_HEADERS.get(col_name, cell.value)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        if cell.value:
            col_max_len[col_idx] = len(str(cell.value))

    num_q_col = None
    status_col = None
    sim_col = None
    for i, c in enumerate(REVERSE_COLUMNS):
        if c == "db_num_questions":
            num_q_col = i + 1
        elif c == "coverage_status":
            status_col = i + 1
        elif c == "similarity_score":
            sim_col = i + 1

    for row_idx in range(2, num_rows + 1):
        rr_idx = row_idx - 2
        rr = reverse_rows[rr_idx] if rr_idx < len(reverse_rows) else None
        hex_value = rr.db_cor_hex if rr else None
        badge = BADGE_COLORS.get(hex_value) if hex_value else None

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.font = _DATA_FONT

            # Track column width
            if cell.value:
                col_max_len[col_idx] = max(col_max_len[col_idx], len(str(cell.value)))

            if col_idx == num_q_col and badge:
                # Question count cell keeps its colored badge
                cell.fill = _get_fill(badge[0])
                cell.font = _BADGE_FONT_CACHE.get(hex_value, _BOLD_FONT)
            elif col_idx == status_col and rr:
                # Coverage status keeps its semantic color (coberto/parcial/não coberto)
                cov_color = COVERAGE_COLORS.get(rr.coverage_status)
                if cov_color:
                    cell.fill = _get_fill(cov_color)
                cell.font = _BOLD_FONT
            else:
                # All other cells get uniform light blue
                cell.fill = _ROW_FILL
                if col_idx == sim_col:
                    cell.font = _BOLD_FONT

    ws.row_dimensions[1].height = 30
    for row_idx in range(2, num_rows + 1):
        ws.row_dimensions[row_idx].height = 25

    # Apply column widths (already computed during loop)
    for col_idx in range(1, num_cols + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(
            max(col_max_len[col_idx] + 4, 15), 55
        )


